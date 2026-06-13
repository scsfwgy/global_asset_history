"""A-share leader stock breakout analysis.

Scans eligible stocks for 6+ consecutive limit-up days,
then analyzes the subsequent pullback and new-high breakthrough.

Data source: AKShare (Sina finance for kline, stock list).
"""

from __future__ import annotations

import json
import logging
import os
import threading
import time
from concurrent.futures import ThreadPoolExecutor, as_completed
from dataclasses import dataclass
from datetime import date, datetime, timezone
from pathlib import Path
from typing import Dict, List, Optional, Tuple

import requests

try:
    import akshare as ak
except ImportError:
    ak = None  # type: ignore

try:
    import pandas as pd
except ImportError:
    pd = None  # type: ignore

logger = logging.getLogger(__name__)

# ---------------------------------------------------------------------------
# Config
# ---------------------------------------------------------------------------

_STOCK_LIST_CACHE_PATH = Path(__file__).resolve().parents[3] / "backend" / "config" / "a_stock_list.json"


@dataclass
class AStockKline:
    date: date
    open: float
    close: float
    high: float
    low: float
    change_pct: float  # daily return %


@dataclass
class BreakoutResult:
    code: str
    name: str
    first_streak_start: str
    first_streak_end: str
    consecutive_limit_up_days: int
    peak_price: float
    peak_date: str
    next_day_limit_down: bool
    limit_down_date: Optional[str]
    pullback_days: int
    bottom_price: float
    bottom_date: str
    breakthrough_days: Optional[int]
    breakthrough_date: Optional[str]
    breakthrough_price: Optional[float]
    new_high: Optional[float]
    new_high_date: Optional[str]


# ---------------------------------------------------------------------------
# Cache
# ---------------------------------------------------------------------------

LEADER_CACHE_TTL = 4 * 60 * 60
_leader_caches: Dict[str, Dict] = {}       # key → result dict
_leader_cache_times: Dict[str, float] = {}  # key → timestamp
_leader_cache_lock = threading.RLock()

# Background scan state — one entry per cache key
_scan_statuses: Dict[str, Dict] = {}  # key → {"status":"scanning"|"done"|"error", ...}
_scan_lock = threading.RLock()


def _get_cache(start_date: str, threshold: float, min_days: int) -> Optional[Dict]:
    """Return cached scan results if valid, else None."""
    ck = _cache_key(start_date, threshold, min_days)
    with _leader_cache_lock:
        entry = _leader_caches.get(ck)
        ts = _leader_cache_times.get(ck, 0)
        if entry is not None and time.time() - ts < LEADER_CACHE_TTL:
            return dict(entry)  # shallow copy to avoid mutation
    return None


def start_background_scan(
    start_date: str, threshold: float, min_days: int,
    workers: int = 10, max_stocks: int = 0,
) -> None:
    """Launch a background scan thread. Returns immediately."""
    ck = _cache_key(start_date, threshold, min_days)

    with _scan_lock:
        existing = _scan_statuses.get(ck, {})
        if existing.get("status") == "scanning":
            logger.info("Background scan already running for key=%s, skipping", ck)
            return
        _scan_statuses[ck] = {"status": "scanning", "key": ck, "started_at": time.time()}

    def _run():
        try:
            logger.info("Background scan started: key=%s", ck)
            run_leader_breakout_scan(
                start_date=start_date,
                threshold=threshold,
                min_consecutive_days=min_days,
                workers=workers,
                force_refresh=True,
                max_stocks=max_stocks,
            )
            with _scan_lock:
                _scan_statuses[ck] = {"status": "done", "key": ck, "done_at": time.time()}
            logger.info("Background scan completed: key=%s", ck)
        except Exception as e:
            logger.exception("Background scan failed: key=%s, error=%s", ck, e)
            with _scan_lock:
                _scan_statuses[ck] = {"status": "error", "key": ck, "error": str(e), "done_at": time.time()}

    t = threading.Thread(target=_run, daemon=True)
    t.start()
    logger.info("Background scan thread started: key=%s", ck)


def get_scan_status(start_date: str = None, threshold: float = None, min_days: int = None) -> Dict:
    """Return scan status for given params, or the first active scan if no params given."""
    with _scan_lock:
        if start_date and threshold is not None and min_days is not None:
            ck = _cache_key(start_date, threshold, min_days)
            return dict(_scan_statuses.get(ck, {"status": "idle"}))
        # Return first non-idle status, or idle
        for s in _scan_statuses.values():
            if s.get("status") != "done":
                return dict(s)
        return {"status": "idle"}


# ---------------------------------------------------------------------------
# Stock list — AKShare + local JSON cache
# ---------------------------------------------------------------------------

def _is_excluded_stock(code: str, name: str) -> bool:
    s = code.strip().upper()
    if s[:3] in ("300", "301", "688"):
        return True
    nm = name.upper()
    if "ST" in nm or "*ST" in nm:
        return True
    return False


def _discover_stocks_akshare() -> List[Dict[str, str]]:
    """Fetch all A-share stocks via AKShare, filter to main board only."""
    if ak is None:
        raise ImportError("akshare未安装，请联系管理员执行: pip install akshare")
    df = ak.stock_info_a_code_name()
    stocks: List[Dict[str, str]] = []
    for _, row in df.iterrows():
        code = str(row["code"]).strip().zfill(6)
        name = str(row["name"]).strip()
        if not _is_excluded_stock(code, name):
            stocks.append({"code": code, "name": name})

    logger.info("AKShare: %d total → %d main board eligible", len(df), len(stocks))

    _STOCK_LIST_CACHE_PATH.parent.mkdir(parents=True, exist_ok=True)
    with open(_STOCK_LIST_CACHE_PATH, "w", encoding="utf-8") as f:
        json.dump(stocks, f, ensure_ascii=False, indent=2)

    return stocks


def _load_cached_stock_list() -> Optional[List[Dict[str, str]]]:
    if not _STOCK_LIST_CACHE_PATH.exists():
        return None
    try:
        with open(_STOCK_LIST_CACHE_PATH, "r", encoding="utf-8") as f:
            data = json.load(f)
        if isinstance(data, list) and len(data) > 100:
            return data
    except (json.JSONDecodeError, IOError) as e:
        logger.warning("Failed to load stock list cache: %s", e)
    return None


def fetch_eligible_stocks() -> List[Dict[str, str]]:
    cached = _load_cached_stock_list()
    if cached:
        logger.info("Using cached stock list: %d stocks", len(cached))
        return cached
    logger.info("No cache, fetching stock list via AKShare...")
    try:
        return _discover_stocks_akshare()
    except Exception as e:
        logger.error("AKShare stock list failed: %s", e)
        return []


# ---------------------------------------------------------------------------
# Kline fetching — two-phase: concurrent HTTP + serial MiniRacer decode
# MiniRacer (V8) is NOT thread-safe, so we download raw data in parallel
# and decode everything in the main thread.
# ---------------------------------------------------------------------------

try:
    import py_mini_racer
    from akshare.stock.cons import zh_sina_a_stock_hist_url, hk_js_decode
except ImportError:
    py_mini_racer = None  # type: ignore
    zh_sina_a_stock_hist_url = ""
    hk_js_decode = ""

# Pre-compile the JS decoder once (thread-safe as long as single MiniRacer instance)
_mini_racer = None
_js_runtime = None  # "mini_racer" | "execjs" | None


def _get_mini_racer():
    """Return a MiniRacer instance, falling back to execjs (Node.js) on platforms
    where py_mini_racer has native library issues (e.g. Apple Silicon)."""
    global _mini_racer, _js_runtime
    if _js_runtime:
        return _mini_racer  # already initialized

    if py_mini_racer is not None and _mini_racer is None:
        try:
            _mini_racer = py_mini_racer.MiniRacer()
            _mini_racer.eval(hk_js_decode)
            _js_runtime = "mini_racer"
            logger.info("JS runtime: py_mini_racer")
            return _mini_racer
        except Exception as e:
            logger.warning("py_mini_racer init failed (%s), falling back to execjs/Node.js", e)
            _mini_racer = None

    # Fallback: use execjs (requires Node.js)
    if _mini_racer is None:
        try:
            import execjs
            _mini_racer = execjs  # not a MiniRacer, but we wrap call() below
        except ImportError:
            raise ImportError(
                "JS 运行时不可用。py_mini_racer 初始化失败且 execjs 未安装。"
                "请安装: pip install py_mini_racer 或 pip install PyExecJS"
            )
        _js_runtime = "execjs"
        logger.info("JS runtime: execjs (Node.js)")
        return _mini_racer

    return _mini_racer


def _js_call(decoder_func_name: str, encoded_str: str) -> list:
    """Call the JS decode function, abstracting over MiniRacer vs execjs."""
    if _js_runtime == "execjs":
        import execjs
        ctx = execjs.compile(hk_js_decode)
        return ctx.call(decoder_func_name, encoded_str)
    else:
        return _get_mini_racer().call(decoder_func_name, encoded_str)


def _code_to_ak_symbol(code: str) -> str:
    """Convert '000001' → 'sz000001', '600000' → 'sh600000'."""
    s = code.strip().upper()
    if s[:3] in ("000", "001", "002", "003", "300", "301"):
        return f"sz{s}"
    return f"sh{s}"


def _fetch_raw_sina_text(code: str) -> Optional[str]:
    """HTTP GET raw kline JS text from Sina. Thread-safe (no MiniRacer)."""
    symbol = _code_to_ak_symbol(code)
    url = zh_sina_a_stock_hist_url.format(symbol)
    try:
        resp = requests.get(url, timeout=15)
        resp.raise_for_status()
        return resp.text
    except Exception as e:
        logger.warning("Sina HTTP failed for %s: %s", code, e)
        return None


def _decode_sina_klines(raw_texts: Dict[str, str]) -> Dict:
    """Decode a batch of raw Sina JS texts into DataFrames. Call from main thread."""
    if pd is None:
        raise ImportError("pandas未安装，请联系管理员执行: pip install pandas")
    _get_mini_racer()  # ensure JS runtime is initialized
    results = {}
    for code, text in raw_texts.items():
        try:
            encoded = text.split("=")[1].split(";")[0].replace('"', "")
            dict_list = _js_call("d", encoded)
            df = pd.DataFrame(dict_list)
            df.index = pd.to_datetime(df["date"], errors="coerce").dt.date
            del df["date"]
            df = df.rename(columns={
                "open": "open", "high": "high", "close": "close",
                "low": "low", "volume": "volume",
            })
            results[code] = df
        except Exception as e:
            logger.warning("Sina decode failed for %s: %s", code, e)
    return results


def _df_to_klines(df, start_date: str) -> List[AStockKline]:
    """Convert a decoded DataFrame to AStockKline list with computed change_pct."""
    result: List[AStockKline] = []
    prev_close: Optional[float] = None

    for idx, row in df.iterrows():
        d = idx if hasattr(idx, "date") else idx
        if hasattr(d, "date"):
            d = d.date()
        try:
            close = float(row["close"])
        except (ValueError, KeyError):
            continue

        chg = 0.0
        if prev_close is not None and prev_close > 0:
            chg = round((close / prev_close - 1) * 100, 2)
        prev_close = close

        result.append(AStockKline(
            date=d,
            open=float(row["open"]),
            close=close,
            high=float(row["high"]),
            low=float(row["low"]),
            change_pct=chg,
        ))

    # Filter by start_date
    sd = datetime.strptime(start_date, "%Y-%m-%d").date()
    return [k for k in result if k.date >= sd]


def fetch_kline_with_details(code: str, start_date: str = "2024-09-01") -> List[AStockKline]:
    """Legacy single-stock fetch (used by tests). Uses the two-phase internally."""
    import pandas as pd
    raw = _fetch_raw_sina_text(code)
    if not raw:
        return []
    decoded = _decode_sina_klines({code: raw})
    df = decoded.get(code)
    if df is None or df.empty:
        return []
    return _df_to_klines(df, start_date)


# ---------------------------------------------------------------------------
# Limit-up streak detection
# ---------------------------------------------------------------------------

def detect_limit_up_streaks(
    klines: List[AStockKline],
    threshold: float = 9.5,
    min_days: int = 6,
) -> List[Tuple[int, int]]:
    streaks: List[Tuple[int, int]] = []
    in_streak = False
    streak_start = 0

    for i, k in enumerate(klines):
        is_limit_up = k.change_pct >= threshold
        if is_limit_up and not in_streak:
            in_streak = True
            streak_start = i
        elif not is_limit_up and in_streak:
            streak_end = i - 1
            if streak_end - streak_start + 1 >= min_days:
                streaks.append((streak_start, streak_end))
            in_streak = False

    if in_streak:
        streak_end = len(klines) - 1
        if streak_end - streak_start + 1 >= min_days:
            streaks.append((streak_start, streak_end))

    return streaks


# ---------------------------------------------------------------------------
# Breakout analysis
# ---------------------------------------------------------------------------

def analyze_breakout(
    klines: List[AStockKline],
    streak_start: int,
    streak_end: int,
) -> Optional[BreakoutResult]:
    """Analyze pullback and new-high breakthrough after a streak."""
    first_streak_start = klines[streak_start]
    first_streak_end = klines[streak_end]
    n = len(klines)

    # 1. Peak: highest close before a clear reversal
    peak_idx = streak_end
    consecutive_down = 0

    for i in range(streak_end + 1, min(n, streak_end + 21)):
        k = klines[i]
        if k.close > klines[peak_idx].close:
            peak_idx = i
        if k.change_pct <= -9.5:
            break
        if k.close < klines[i - 1].close:
            consecutive_down += 1
            if consecutive_down >= 2:
                break
        else:
            consecutive_down = 0

    peak = klines[peak_idx]

    # 2. Next-day limit-down
    next_day_limit_down = False
    limit_down_date: Optional[str] = None
    if streak_end + 1 < n:
        nd = klines[streak_end + 1]
        if nd.change_pct <= -9.5:
            next_day_limit_down = True
            limit_down_date = nd.date.isoformat()

    # 3. Bottom + Breakthrough
    bottom_start = peak_idx + 1
    if bottom_start >= n:
        return BreakoutResult(
            code="", name="",
            first_streak_start=first_streak_start.date.isoformat(),
            first_streak_end=first_streak_end.date.isoformat(),
            consecutive_limit_up_days=streak_end - streak_start + 1,
            peak_price=round(peak.close, 2),
            peak_date=peak.date.isoformat(),
            next_day_limit_down=next_day_limit_down,
            limit_down_date=limit_down_date,
            pullback_days=0,
            bottom_price=round(peak.close, 2),
            bottom_date=peak.date.isoformat(),
            breakthrough_days=None, breakthrough_date=None, breakthrough_price=None,
            new_high=None, new_high_date=None,
        )

    true_bottom_idx = bottom_start
    breakthrough_idx: Optional[int] = None

    for i in range(bottom_start, n):
        if breakthrough_idx is None and klines[i].close < klines[true_bottom_idx].close:
            true_bottom_idx = i
        if breakthrough_idx is None and klines[i].close > peak.close:
            breakthrough_idx = i

    bottom = klines[true_bottom_idx]
    pullback_days = max(0, true_bottom_idx - peak_idx)

    bt_days: Optional[int] = None
    bt_date: Optional[str] = None
    bt_price: Optional[float] = None
    if breakthrough_idx is not None:
        bt_days = breakthrough_idx - true_bottom_idx
        bt_date = klines[breakthrough_idx].date.isoformat()
        bt_price = round(klines[breakthrough_idx].close, 2)

    new_high: Optional[float] = None
    new_high_date: Optional[str] = None
    if breakthrough_idx is not None:
        nh_start = breakthrough_idx
        if nh_start < n:
            nh_idx = nh_start
            for i in range(nh_start, n):
                if klines[i].close > klines[nh_idx].close:
                    nh_idx = i
            new_high = round(klines[nh_idx].close, 2)
            new_high_date = klines[nh_idx].date.isoformat()

    return BreakoutResult(
        code="", name="",
        first_streak_start=first_streak_start.date.isoformat(),
        first_streak_end=first_streak_end.date.isoformat(),
        consecutive_limit_up_days=streak_end - streak_start + 1,
        peak_price=round(peak.close, 2),
        peak_date=peak.date.isoformat(),
        next_day_limit_down=next_day_limit_down,
        limit_down_date=limit_down_date,
        pullback_days=pullback_days,
        bottom_price=round(bottom.close, 2),
        bottom_date=bottom.date.isoformat(),
        breakthrough_days=bt_days, breakthrough_date=bt_date, breakthrough_price=bt_price,
        new_high=new_high, new_high_date=new_high_date,
    )


# ---------------------------------------------------------------------------
# Orchestration
# ---------------------------------------------------------------------------

def _process_one_stock_from_df(
    code: str, name: str, df, start_date: str, threshold: float, min_days: int,
) -> Optional[Dict]:
    """Analyze a single stock from its already-decoded DataFrame."""
    klines = _df_to_klines(df, "2024-09-01")

    analysis_start = datetime.strptime(start_date, "%Y-%m-%d").date()
    filtered = [k for k in klines if k.date >= analysis_start]
    if len(filtered) < min_days:
        return None

    streaks = detect_limit_up_streaks(filtered, threshold=threshold, min_days=min_days)
    if not streaks:
        return None

    result = analyze_breakout(filtered, streaks[0][0], streaks[0][1])
    result.code = code
    result.name = name

    # Log qualifying stocks
    bt_str = ""
    if result.breakthrough_days is not None:
        bt_str = f" 突破{result.breakthrough_days}d→{result.new_high}"
    logger.info(
        "  ✅ %s(%s) 涨停%dd %s~%s 高峰%.2f 底%.2f 回调%dd%s",
        name, code,
        result.consecutive_limit_up_days,
        result.first_streak_start, result.first_streak_end,
        result.peak_price, result.bottom_price, result.pullback_days,
        bt_str,
    )

    return {
        "code": result.code, "name": result.name,
        "first_streak_start": result.first_streak_start,
        "first_streak_end": result.first_streak_end,
        "consecutive_limit_up_days": result.consecutive_limit_up_days,
        "peak_price": result.peak_price, "peak_date": result.peak_date,
        "next_day_limit_down": result.next_day_limit_down,
        "limit_down_date": result.limit_down_date,
        "pullback_days": result.pullback_days,
        "bottom_price": result.bottom_price, "bottom_date": result.bottom_date,
        "breakthrough_days": result.breakthrough_days,
        "breakthrough_date": result.breakthrough_date,
        "breakthrough_price": result.breakthrough_price,
        "new_high": result.new_high, "new_high_date": result.new_high_date,
    }


def _cache_key(start_date: str, threshold: float, min_days: int) -> str:
    return f"{start_date}|{threshold}|{min_days}"


def run_leader_breakout_scan(
    start_date: str = "2024-09-30",
    threshold: float = 9.5,
    min_consecutive_days: int = 6,
    workers: int = 10,
    force_refresh: bool = False,
    max_stocks: int = 0,
) -> Dict:
    """Two-phase scan:
    Phase 1: concurrent HTTP fetch of raw Sina kline text (thread-safe)
    Phase 2: serial MiniRacer decode in main thread
    Phase 3: concurrent streak detection + breakout analysis
    """
    global _leader_caches, _leader_cache_times
    ck = _cache_key(start_date, threshold, min_consecutive_days)
    with _leader_cache_lock:
        if not force_refresh:
            entry = _leader_caches.get(ck)
            ts = _leader_cache_times.get(ck, 0)
            if entry is not None and time.time() - ts < LEADER_CACHE_TTL:
                logger.info("Returning cached leader breakout results")
                return entry

    start_ts = time.time()

    all_stocks = fetch_eligible_stocks()
    if max_stocks > 0:
        all_stocks = all_stocks[:max_stocks]
    total = len(all_stocks)

    # ── Phase 1: concurrent HTTP fetch (no MiniRacer, fully thread-safe) ──
    raw_texts: Dict[str, str] = {}
    fetched = 0
    fetch_errors = 0

    actual_workers = min(workers, max(1, total))
    with ThreadPoolExecutor(max_workers=actual_workers) as executor:
        future_to_stock = {
            executor.submit(_fetch_raw_sina_text, s["code"]): s
            for s in all_stocks
        }
        for future in as_completed(future_to_stock):
            fetched += 1
            stock = future_to_stock[future]
            try:
                text = future.result()
                if text:
                    raw_texts[stock["code"]] = text
                else:
                    fetch_errors += 1
            except Exception:
                fetch_errors += 1

    logger.info("Phase 1 done: %d fetched, %d errors, %.1fs",
                len(raw_texts), fetch_errors, time.time() - start_ts)

    # ── Phase 2: serial MiniRacer decode (main thread only) ──
    decoded = _decode_sina_klines(raw_texts)
    logger.info("Phase 2 done: %d decoded, %.1fs", len(decoded), time.time() - start_ts)

    # ── Phase 3: concurrent analysis ──
    results: List[Dict] = []
    scan_errors = 0
    analyzed = 0
    to_analyze = len(decoded)

    stock_lookup = {s["code"]: s["name"] for s in all_stocks}

    with ThreadPoolExecutor(max_workers=actual_workers) as executor:
        future_to_code = {}
        for code, df in decoded.items():
            name = stock_lookup.get(code, code)
            future = executor.submit(
                _process_one_stock_from_df,
                code, name, df, start_date, threshold, min_consecutive_days,
            )
            future_to_code[future] = code

        for future in as_completed(future_to_code):
            analyzed += 1
            if analyzed % 500 == 0:
                logger.info("  Phase 3 progress: %d/%d analyzed, %d qualified",
                           analyzed, to_analyze, len(results))
            try:
                r = future.result()
                if r is not None:
                    results.append(r)
            except Exception:
                scan_errors += 1

    results.sort(key=lambda r: r["first_streak_start"], reverse=True)

    elapsed = round(time.time() - start_ts, 1)
    logger.info("Scan done: %d stocks, %d qualified, %.1fs", total, len(results), elapsed)

    qualified = len(results)
    recovered = sum(1 for r in results if r["breakthrough_days"] is not None)
    bt_list = [r["breakthrough_days"] for r in results if r["breakthrough_days"] is not None]
    avg_bt = round(sum(bt_list) / len(bt_list), 1) if bt_list else None
    avg_pb = round(sum(r["pullback_days"] for r in results) / qualified, 1) if qualified else None

    output = {
        "_key": ck, "_cached_at": datetime.now(timezone.utc).isoformat(),
        "summary": {
            "total_stocks_scanned": total, "qualified": qualified,
            "recovered": recovered, "not_recovered": qualified - recovered,
            "avg_pullback_days": avg_pb, "avg_breakthrough_days": avg_bt,
            "scan_time_seconds": elapsed,
        },
        "stocks": results,
    }

    with _leader_cache_lock:
        _leader_caches[ck] = output
        _leader_cache_times[ck] = time.time()
        # Cleanup stale entries (> 2x TTL) to prevent unbounded growth
        stale = [k for k, t in _leader_cache_times.items() if time.time() - t > LEADER_CACHE_TTL * 2]
        for k in stale:
            _leader_caches.pop(k, None)
            _leader_cache_times.pop(k, None)

    return output


def export_leader_breakout_excel(scan_result: Dict) -> bytes:
    """Generate an Excel file from leader breakout scan results."""
    import io
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = Workbook()
    ws = wb.active
    ws.title = "A股龙头股回调新高统计"

    # Header styling
    header_font = Font(name="Microsoft YaHei", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="0071E3", end_color="0071E3", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )
    cell_font = Font(name="Microsoft YaHei", size=10)
    cell_align = Alignment(horizontal="center", vertical="center")
    green_font = Font(name="Microsoft YaHei", size=10, color="30D158")
    red_font = Font(name="Microsoft YaHei", size=10, color="FF453A")

    # Summary row
    s = scan_result.get("summary", {})
    ws.merge_cells("A1:I1")
    ws["A1"] = (
        f"扫描股票: {s.get('total_stocks_scanned', 0)} | "
        f"符合条件: {s.get('qualified', 0)} | "
        f"已突破: {s.get('recovered', 0)} | "
        f"未突破: {s.get('not_recovered', 0)} | "
        f"平均回调: {s.get('avg_pullback_days', '—')}天 | "
        f"平均突破: {s.get('avg_breakthrough_days', '—')}天 | "
        f"耗时: {s.get('scan_time_seconds', 0)}秒"
    )
    ws["A1"].font = Font(name="Microsoft YaHei", size=10, bold=True)

    # Headers
    headers = [
        "股票名称", "股票代码", "首次涨停日期", "涨停天数",
        "高峰价格", "次日跌停", "回调天数", "低点价格",
        "突破天数", "新高价格",
    ]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    # Data rows
    stocks = scan_result.get("stocks", [])
    for row_idx, r in enumerate(stocks, 4):
        values = [
            r["name"], r["code"], r["first_streak_start"],
            r["consecutive_limit_up_days"], r["peak_price"],
            "是" if r["next_day_limit_down"] else "否",
            r["pullback_days"], r["bottom_price"],
            r["breakthrough_days"] if r["breakthrough_days"] is not None else "—",
            r["new_high"] if r["new_high"] is not None else "—",
        ]
        for col, val in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col, value=val)
            cell.font = cell_font
            cell.alignment = cell_align
            cell.border = thin_border
            # Color breakthrough column green if recovered
            if col == 9 and r["breakthrough_days"] is not None:
                cell.font = green_font
            if col == 10 and r["new_high"] is not None:
                cell.font = green_font

    # Column widths
    widths = [14, 10, 14, 8, 10, 8, 8, 10, 8, 10]
    for col, w in enumerate(widths, 1):
        ws.column_dimensions[ws.cell(row=3, column=col).column_letter].width = w

    # Freeze header
    ws.freeze_panes = "A4"

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
